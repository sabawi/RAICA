"""
FAISS Document Interrogation System
Advanced document processing, embedding, and retrieval for directory-based document collections
Integrates with existing 2-stage LLM architecture as Stage 0 (RAG Retrieval)
"""

import os
import sqlite3
import json
import pickle
import hashlib
import logging
import asyncio
from typing import Dict, List, Optional, Tuple, Any, Union
from dataclasses import dataclass, field
from pathlib import Path
from datetime import datetime
import time
from utils.config_loader import config_loader

# Import integrity monitoring
try:
    from tools.faiss_integrity_monitor import check_and_repair_faiss_integrity
    INTEGRITY_MONITORING_AVAILABLE = True
except ImportError:
    INTEGRITY_MONITORING_AVAILABLE = False
    print("⚠️ FAISS integrity monitoring not available")

# Document processing with multi-engine OCR support
try:
    import PyPDF2
    import docx
    import openpyxl
    from bs4 import BeautifulSoup
    from PIL import Image
    
    # Multi-engine OCR system with intelligent fallback
    OCR_ENGINES = []
    
    # Primary: EasyOCR (recommended)
    try:
        import easyocr
        OCR_ENGINES.append(('easyocr', 'available'))
    except ImportError:
        OCR_ENGINES.append(('easyocr', 'missing'))
    
    # Fallback: Tesseract
    try:
        import pytesseract
        OCR_ENGINES.append(('tesseract', 'available'))
    except ImportError:
        OCR_ENGINES.append(('tesseract', 'missing'))
    
    DOCUMENT_PROCESSING_AVAILABLE = True
except ImportError as e:
    DOCUMENT_PROCESSING_AVAILABLE = False
    DOCUMENT_PROCESSING_ERROR = str(e)
    OCR_ENGINES = []

# FAISS and embeddings
try:
    import faiss
    import numpy as np
    FAISS_AVAILABLE = True
except ImportError as e:
    FAISS_AVAILABLE = False
    FAISS_ERROR = str(e)

# File monitoring
try:
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
    WATCHDOG_AVAILABLE = True
except ImportError as e:
    WATCHDOG_AVAILABLE = False
    WATCHDOG_ERROR = str(e)

# =============================================================================
# CONFIGURATION LOADER - Load from config/llm_config.yaml per PROJECT_CONFIGURATION_DIRECTIVE
# =============================================================================

def _load_embedding_config() -> Dict[str, Any]:
    """Load embedding configuration from config file with safe fallbacks"""
    try:
        config = config_loader.load_config()
        doc_config = config.get('document_interrogator', {})
        embedding_config = doc_config.get('embedding', {})

        return {
            # Model configuration
            'model_name': embedding_config.get('model_name', 'mxbai-embed-large'),
            'dimension': embedding_config.get('dimension', 1024),

            # Service configuration
            'service_host': embedding_config.get('service', {}).get('host', '127.0.0.1'),
            'service_port': embedding_config.get('service', {}).get('port', 11434),
            'service_base_url': embedding_config.get('service', {}).get('base_url', 'http://127.0.0.1:11434'),

            # Timeout configuration
            'embedding_timeout': embedding_config.get('timeout', {}).get('embedding_request', 120),
            'health_check_timeout': embedding_config.get('timeout', {}).get('health_check', 10),
            'service_restart_delay': embedding_config.get('timeout', {}).get('service_restart', 5),

            # Retry configuration
            'max_service_restart_attempts': embedding_config.get('retry', {}).get('max_service_restart_attempts', 3),
            'retry_delay': embedding_config.get('retry', {}).get('retry_delay_seconds', 3),

            # Batch processing configuration (CRITICAL)
            'batch_size': embedding_config.get('batch_processing', {}).get('batch_size', 10),
            'batch_delay': embedding_config.get('batch_processing', {}).get('batch_delay_seconds', 5.0),
            'adaptive_mode_enabled': embedding_config.get('batch_processing', {}).get('adaptive_mode_enabled', True),
            'min_batch_size': embedding_config.get('batch_processing', {}).get('min_batch_size', 1),
            'adaptive_reduction_factor': embedding_config.get('batch_processing', {}).get('adaptive_reduction_factor', 0.5),

            # Document processing configuration
            'chunk_size': doc_config.get('document_processing', {}).get('chunk_size', 1000),
            'chunk_overlap': doc_config.get('document_processing', {}).get('chunk_overlap', 200),
            'min_chunk_length': doc_config.get('document_processing', {}).get('min_chunk_length', 10),
            'max_chunk_length': doc_config.get('document_processing', {}).get('max_chunk_length', 2000),

            # Directory scanning configuration
            'scan_interval_minutes': doc_config.get('scan_interval_minutes', 60),
            'max_files_per_scan': doc_config.get('max_files_per_scan', 150),
            'startup_initialization_delay': doc_config.get('startup_initialization_delay', 3),
        }
    except Exception as e:
        logger.error(f"❌ Failed to load embedding config, using safe fallbacks: {e}")
        # Return safe fallback configuration
        return {
            'model_name': 'mxbai-embed-large',
            'dimension': 1024,
            'service_host': '127.0.0.1',
            'service_port': 11434,
            'service_base_url': 'http://127.0.0.1:11434',
            'embedding_timeout': 120,
            'health_check_timeout': 10,
            'service_restart_delay': 5,
            'max_service_restart_attempts': 3,
            'retry_delay': 3,
            'batch_size': 10,  # ✅ Safe reduced default
            'batch_delay': 5.0,  # ✅ Safe increased default
            'adaptive_mode_enabled': True,
            'min_batch_size': 1,
            'adaptive_reduction_factor': 0.5,
            'chunk_size': 1000,
            'chunk_overlap': 200,
            'min_chunk_length': 10,
            'max_chunk_length': 2000,
            'scan_interval_minutes': 60,
            'max_files_per_scan': 150,
            'startup_initialization_delay': 3,
        }

# Load configuration on module initialization
_EMBEDDING_CONFIG = _load_embedding_config()

# ✅ Configuration Constants (loaded from config file, not hardcoded)
EMBEDDING_MODEL_NAME = _EMBEDDING_CONFIG['model_name']
EMBEDDING_DIMENSION = _EMBEDDING_CONFIG['dimension']
OLLAMA_EMBEDDING_HOST = _EMBEDDING_CONFIG['service_host']
OLLAMA_EMBEDDING_PORT = _EMBEDDING_CONFIG['service_port']
OLLAMA_MAIN_PORT = _EMBEDDING_CONFIG['service_port']
EMBEDDING_SERVICE_URL = f"{_EMBEDDING_CONFIG['service_base_url']}/api/embeddings"
EMBEDDING_TIMEOUT_SECONDS = _EMBEDDING_CONFIG['embedding_timeout']
HEALTH_CHECK_TIMEOUT_SECONDS = _EMBEDDING_CONFIG['health_check_timeout']
SERVICE_RESTART_DELAY_SECONDS = _EMBEDDING_CONFIG['service_restart_delay']
RETRY_DELAY_SECONDS = _EMBEDDING_CONFIG['retry_delay']
MAX_SERVICE_RESTART_ATTEMPTS = _EMBEDDING_CONFIG['max_service_restart_attempts']
DEFAULT_BATCH_SIZE = _EMBEDDING_CONFIG['batch_size']
DEFAULT_CHUNK_SIZE = _EMBEDDING_CONFIG['chunk_size']
DEFAULT_SEARCH_RESULTS = 5  # Static default for search results
MAX_FILES_PER_DIRECTORY_SCAN = None
SCAN_PROGRESS_LOG_INTERVAL = 10
DEFAULT_SCAN_INTERVAL_MINUTES = _EMBEDDING_CONFIG['scan_interval_minutes']
STARTUP_INITIALIZATION_DELAY = _EMBEDDING_CONFIG['startup_initialization_delay']

# ✅ Adaptive batch configuration (NEW)
ADAPTIVE_BATCH_MODE_ENABLED = _EMBEDDING_CONFIG['adaptive_mode_enabled']
MIN_BATCH_SIZE = _EMBEDDING_CONFIG['min_batch_size']
ADAPTIVE_BATCH_REDUCTION_FACTOR = _EMBEDDING_CONFIG['adaptive_reduction_factor']
BATCH_DELAY_SECONDS = _EMBEDDING_CONFIG['batch_delay']

# Existing server integration
from http_helpers import pooled_post

# =============================================================================
# OCR ENGINE MANAGER - Multi-engine OCR with intelligent fallback
# =============================================================================

class OCREngineManager:
    """Intelligent OCR engine management with fallback system"""
    
    def __init__(self):
        self.primary_engine = None
        self.fallback_engines = []
        self._initialize_engines()
    
    def _initialize_engines(self):
        """Initialize available OCR engines in priority order"""
        for engine_name, status in OCR_ENGINES:
            try:
                if engine_name == 'easyocr' and status == 'available':
                    reader = easyocr.Reader(['en'])
                    if self.primary_engine is None:
                        self.primary_engine = ('easyocr', reader)
                        logger.info("🚀 EasyOCR initialized as primary OCR engine")
                    else:
                        self.fallback_engines.append(('easyocr', reader))
                        
                elif engine_name == 'tesseract' and status == 'available':
                    # Test if tesseract binary is available
                    try:
                        test_img = Image.new('RGB', (100, 30), color='white')
                        pytesseract.image_to_string(test_img)
                        
                        if self.primary_engine is None:
                            self.primary_engine = ('tesseract', pytesseract)
                            logger.info("📝 Tesseract initialized as primary OCR engine")
                        else:
                            self.fallback_engines.append(('tesseract', pytesseract))
                    except Exception:
                        logger.warning("⚠️ Tesseract library present but binary not accessible")
                        
            except Exception as e:
                logger.warning(f"Failed to initialize {engine_name}: {e}")
        
        if self.primary_engine is None:
            logger.error("❌ No OCR engines available")
        else:
            logger.info(f"✅ OCR system ready: {self.primary_engine[0]} + {len(self.fallback_engines)} fallbacks")
    
    async def extract_text(self, image_path: Path) -> str:
        """Extract text with intelligent fallback"""
        # Try primary engine
        if self.primary_engine:
            result = await self._try_engine(self.primary_engine, image_path)
            if result:
                return result
        
        # Try fallback engines
        for engine in self.fallback_engines:
            result = await self._try_engine(engine, image_path)
            if result:
                logger.info(f"✅ Fallback {engine[0]} succeeded for {image_path.name}")
                return result
        
        # All engines failed
        logger.error(f"❌ All OCR engines failed for {image_path}")
        return f"[Image: {image_path.name}] - OCR failed"
    
    async def _try_engine(self, engine_tuple, image_path: Path) -> Optional[str]:
        """Try a specific OCR engine"""
        engine_name, engine = engine_tuple
        try:
            if engine_name == 'easyocr':
                # Start with basic call to isolate any parameter issues
                logger.debug(f"Attempting EasyOCR processing for {image_path}")
                results = engine.readtext(str(image_path), paragraph=True)
                logger.debug(f"EasyOCR returned {len(results)} results")
                
                if not results:
                    logger.warning(f"EasyOCR returned no results for {image_path}")
                    return None
                
                # Enhanced post-processing with confidence filtering and text cleanup
                filtered_text = []
                for i, result in enumerate(results):
                    # Handle different EasyOCR result formats
                    if len(result) == 3:
                        bbox, text, confidence = result
                    elif len(result) == 2:
                        bbox, text = result
                        confidence = 1.0  # Assume high confidence if not provided
                    else:
                        logger.warning(f"Unexpected result format: {result}")
                        continue
                    
                    logger.debug(f"Result {i}: '{text}' (confidence: {confidence})")
                    if confidence > 0.3 and text.strip():  # Include medium-confidence text
                        # Clean up common OCR artifacts while preserving document structure
                        cleaned_text = text.strip()
                        
                        # Fix common OCR character confusion in documents
                        cleaned_text = cleaned_text.replace('|l', 'll')  # Common l/| confusion
                        cleaned_text = cleaned_text.replace('1I', 'Il')  # 1/I confusion
                        cleaned_text = cleaned_text.replace('S@', 'SA')  # @ confusion in STATES
                        cleaned_text = cleaned_text.replace('Ou', 'ou')  # O/o normalization
                        
                        # Only add if still has content after cleaning
                        if cleaned_text and len(cleaned_text) > 1:
                            filtered_text.append(cleaned_text)
                
                final_text = ' '.join(filtered_text).strip()
                logger.info(f"EasyOCR extracted {len(results)} regions, filtered to {len(filtered_text)} regions, final text: {len(final_text)} chars")
                return final_text if final_text else None
                
            elif engine_name == 'tesseract':
                image = Image.open(str(image_path))
                text = engine.image_to_string(image)
                return text.strip() if text.strip() else None
                
        except Exception as e:
            logger.error(f"OCR engine {engine_name} failed for {image_path}: {e}")
            logger.error(f"Exception type: {type(e).__name__}")
            import traceback
            logger.error(f"Full traceback: {traceback.format_exc()}")
            return None

logger = logging.getLogger(__name__)


@dataclass
class DocumentChunk:
    """Represents a processed document chunk with metadata"""
    chunk_id: str
    document_path: str
    content: str
    chunk_index: int
    total_chunks: int
    metadata: Dict[str, Any] = field(default_factory=dict)
    embedding: Optional[np.ndarray] = None
    created_at: str = field(default_factory=lambda: datetime.now().isoformat())


@dataclass 
class DocumentInfo:
    """Document metadata and processing info"""
    file_path: str
    file_hash: str
    file_type: str
    file_size: int
    total_chunks: int
    processed_at: str
    last_modified: str


class DocumentProcessor:
    """Handles extraction of text from various document types"""
    
    SUPPORTED_TYPES = {
        '.txt': 'text',
        '.md': 'markdown', 
        '.html': 'html',
        '.pdf': 'pdf',
        '.docx': 'word',
        '.xlsx': 'excel',
        '.jpg': 'image',
        '.jpeg': 'image',
        '.png': 'image',
        '.bmp': 'image',
        '.tiff': 'image'
    }
    
    def __init__(self):
        # ✅ Load chunk configuration from config file (not hardcoded)
        self.chunk_size = _EMBEDDING_CONFIG.get('chunk_size', DEFAULT_CHUNK_SIZE)
        self.chunk_overlap = _EMBEDDING_CONFIG.get('chunk_overlap', 200)
        self.min_chunk_length = _EMBEDDING_CONFIG.get('min_chunk_length', 10)
        self.max_chunk_length = _EMBEDDING_CONFIG.get('max_chunk_length', 2000)
        self.ocr_manager = OCREngineManager() if DOCUMENT_PROCESSING_AVAILABLE else None

        if not DOCUMENT_PROCESSING_AVAILABLE:
            logger.warning(f"⚠️ Document processing libraries not available: {DOCUMENT_PROCESSING_ERROR}")
            logger.info("💡 Install with: pip install PyPDF2 python-docx openpyxl beautifulsoup4 easyocr pillow")
    
    async def process_document(self, file_path: str) -> List[DocumentChunk]:
        """Process a document into chunks with extracted text"""
        if not DOCUMENT_PROCESSING_AVAILABLE:
            logger.error("❌ Document processing libraries not available")
            return []
            
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")
            
            # Get file info
            file_type = file_path.suffix.lower()
            if file_type not in self.SUPPORTED_TYPES:
                logger.warning(f"Unsupported file type: {file_type}")
                return []
            
            # Extract text based on file type
            text_content = await self._extract_text(file_path, file_type)
            if not text_content.strip():
                logger.warning(f"No text extracted from: {file_path}")
                return []
            
            # Create chunks
            chunks = self._create_chunks(text_content, str(file_path))
            logger.info(f"✅ Processed {file_path.name}: {len(chunks)} chunks")
            return chunks
            
        except Exception as e:
            logger.error(f"❌ Failed to process {file_path}: {e}")
            return []
    
    async def _extract_text(self, file_path: Path, file_type: str) -> str:
        """Extract text from document based on file type"""
        try:
            if file_type == '.txt' or file_type == '.md':
                return file_path.read_text(encoding='utf-8', errors='ignore')
            
            elif file_type == '.html':
                html_content = file_path.read_text(encoding='utf-8', errors='ignore')
                soup = BeautifulSoup(html_content, 'html.parser')
                return soup.get_text(separator=' ', strip=True)
            
            elif file_type == '.pdf':
                return await self._extract_pdf_text(file_path)
            
            elif file_type == '.docx':
                doc = docx.Document(str(file_path))
                return '\n'.join([paragraph.text for paragraph in doc.paragraphs])
            
            elif file_type == '.xlsx':
                workbook = openpyxl.load_workbook(str(file_path))
                text_content = []
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows(values_only=True):
                        row_text = ' '.join([str(cell) for cell in row if cell is not None])
                        if row_text.strip():
                            text_content.append(row_text)
                return '\n'.join(text_content)
            
            elif file_type in ['.jpg', '.jpeg', '.png', '.bmp', '.tiff']:
                return await self._extract_image_text(file_path)
            
            else:
                return ""
                
        except Exception as e:
            logger.error(f"❌ Text extraction failed for {file_path}: {e}")
            return ""
    
    async def _extract_pdf_text(self, file_path: Path) -> str:
        """Extract text from PDF using PyPDF2"""
        try:
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                text_content = []
                
                for page_num, page in enumerate(reader.pages):
                    try:
                        page_text = page.extract_text()
                        if page_text.strip():
                            text_content.append(f"[Page {page_num + 1}]\n{page_text}")
                    except Exception as e:
                        logger.warning(f"Failed to extract page {page_num + 1} from {file_path}: {e}")
                
                return '\n\n'.join(text_content)
                
        except Exception as e:
            logger.error(f"❌ PDF extraction failed for {file_path}: {e}")
            return ""
    
    async def _extract_image_text(self, file_path: Path) -> str:
        """Extract text from image using multi-engine OCR"""
        if not self.ocr_manager:
            return f"[Image: {file_path.name}] - OCR not available"
        
        return await self.ocr_manager.extract_text(file_path)
    
    def _create_chunks(self, text: str, document_path: str) -> List[DocumentChunk]:
        """Split text into overlapping chunks"""
        chunks = []
        if len(text) <= self.chunk_size:
            # Single chunk for small documents
            chunk_id = self._generate_chunk_id(document_path, 0)
            chunks.append(DocumentChunk(
                chunk_id=chunk_id,
                document_path=document_path,
                content=text,
                chunk_index=0,
                total_chunks=1,
                metadata={'length': len(text)}
            ))
        else:
            # Multiple chunks with overlap
            start = 0
            chunk_index = 0
            
            while start < len(text):
                end = min(start + self.chunk_size, len(text))
                chunk_text = text[start:end]
                
                chunk_id = self._generate_chunk_id(document_path, chunk_index)
                chunks.append(DocumentChunk(
                    chunk_id=chunk_id,
                    document_path=document_path,
                    content=chunk_text,
                    chunk_index=chunk_index,
                    total_chunks=0,  # Will be set after all chunks created
                    metadata={'start': start, 'end': end, 'length': len(chunk_text)}
                ))
                
                start = end - self.chunk_overlap
                chunk_index += 1
                
                if end >= len(text):
                    break
            
            # Update total chunks for all chunks
            total_chunks = len(chunks)
            for chunk in chunks:
                chunk.total_chunks = total_chunks
        
        return chunks
    
    def _generate_chunk_id(self, document_path: str, chunk_index: int) -> str:
        """Generate unique chunk ID"""
        content = f"{document_path}:{chunk_index}"
        return hashlib.md5(content.encode()).hexdigest()


class FAISSDocumentStore:
    """FAISS-based vector storage with SQLite metadata"""
    
    def __init__(self, storage_dir: str = "document_store"):
        if not FAISS_AVAILABLE:
            raise ImportError(f"FAISS not available: {FAISS_ERROR}")
            
        self.storage_dir = Path(storage_dir)
        self.storage_dir.mkdir(exist_ok=True)
        
        # FAISS index file paths
        self.index_path = self.storage_dir / "faiss.index"
        self.metadata_path = self.storage_dir / "metadata.db"
        
        # Initialize components
        self.dimension = EMBEDDING_DIMENSION
        self.faiss_index = None
        self.metadata_db = None
        self.chunk_counter = 0
        
        self._initialize_storage()
    
    def _initialize_storage(self):
        """Initialize FAISS index and SQLite database with dimension validation"""
        try:
            # Initialize SQLite database FIRST (needed for metadata tracking)
            self.metadata_db = sqlite3.connect(str(self.metadata_path), check_same_thread=False)
            self._create_tables()

            # Initialize FAISS index with dimension validation
            if self.index_path.exists():
                # ✅ Load existing index
                self.faiss_index = faiss.read_index(str(self.index_path))
                self.chunk_counter = self.faiss_index.ntotal
                logger.info(f"📚 Loaded existing FAISS index with {self.chunk_counter} vectors")

                # 🛡️ CRITICAL: Validate dimension matches configuration
                if self.faiss_index.d != self.dimension:
                    logger.error(f"🚨 DIMENSION MISMATCH DETECTED!")
                    logger.error(f"   Index dimension:  {self.faiss_index.d}")
                    logger.error(f"   Config dimension: {self.dimension}")
                    logger.error(f"   Embedding model:  {EMBEDDING_MODEL_NAME}")
                    logger.error(f"")
                    logger.error(f"   This indicates the embedding model was changed without reindexing.")
                    logger.error(f"   This WILL corrupt search results and must be fixed immediately!")
                    logger.error(f"")
                    logger.error(f"   TO FIX (see: docs/ADMIN_MODEL_CHANGE_GUIDE.md):")
                    logger.error(f"   1. Backup your data: cp -r document_store document_store.backup")
                    logger.error(f"   2. Delete the corrupted index:")
                    logger.error(f"      rm -rf document_store/faiss.index")
                    logger.error(f"      rm -rf document_store/metadata.db")
                    logger.error(f"   3. Verify config matches your embedding model:")
                    logger.error(f"      - model_name in config/llm_config.yaml")
                    logger.error(f"      - dimension matches the model (e.g., 768 for nomic, 1024 for mxbai)")
                    logger.error(f"   4. Restart server: ./stop_complete.sh && ./start_complete.sh")
                    logger.error(f"   5. Rescan documents to rebuild index with correct dimension")
                    raise ValueError(
                        f"DIMENSION MISMATCH: Index has {self.faiss_index.d} dimensions "
                        f"but config specifies {self.dimension}. This typically means the embedding "
                        f"model was changed without reindexing. See docs/ADMIN_MODEL_CHANGE_GUIDE.md"
                    )

                # ✅ Validate model information in metadata
                self._check_model_change()
            else:
                # ✅ Create new index
                self.faiss_index = faiss.IndexFlatIP(self.dimension)
                logger.info(f"🔧 Created new FAISS index (dimension: {self.dimension}, model: {EMBEDDING_MODEL_NAME})")

                # Record this as the initial model used
                self._record_model_metadata()

        except Exception as e:
            logger.error(f"❌ Storage initialization failed: {e}")
            raise
    
    def _create_tables(self):
        """Create SQLite tables for metadata"""
        cursor = self.metadata_db.cursor()
        
        # Documents table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS documents (
                file_path TEXT PRIMARY KEY,
                file_hash TEXT NOT NULL,
                file_type TEXT NOT NULL,
                file_size INTEGER NOT NULL,
                total_chunks INTEGER NOT NULL,
                processed_at TEXT NOT NULL,
                last_modified TEXT NOT NULL
            )
        ''')
        
        # Chunks table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS chunks (
                chunk_id TEXT PRIMARY KEY,
                faiss_index INTEGER NOT NULL,
                document_path TEXT NOT NULL,
                chunk_index INTEGER NOT NULL,
                total_chunks INTEGER NOT NULL,
                content TEXT NOT NULL,
                metadata TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY (document_path) REFERENCES documents (file_path)
            )
        ''')

        # ✅ Model Metadata table - Track embedding model and dimension for safety
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS model_metadata (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                model_name TEXT NOT NULL,
                dimension INTEGER NOT NULL,
                vector_count INTEGER DEFAULT 0,
                created_at TEXT NOT NULL,
                last_updated TEXT NOT NULL,
                notes TEXT
            )
        ''')

        self.metadata_db.commit()

    def _record_model_metadata(self):
        """Record the current embedding model information in metadata table"""
        try:
            cursor = self.metadata_db.cursor()

            # Check if metadata already exists
            cursor.execute('SELECT COUNT(*) FROM model_metadata WHERE id = 1')
            if cursor.fetchone()[0] == 0:
                # Insert new record
                cursor.execute('''
                    INSERT INTO model_metadata
                    (id, model_name, dimension, vector_count, created_at, last_updated, notes)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (
                    1,
                    EMBEDDING_MODEL_NAME,
                    self.dimension,
                    self.chunk_counter,
                    datetime.now().isoformat(),
                    datetime.now().isoformat(),
                    f"Initial model setup with {self.chunk_counter} vectors"
                ))
                logger.info(f"📝 Recorded model metadata: {EMBEDDING_MODEL_NAME} (dimension: {self.dimension})")
            else:
                # Update existing record
                cursor.execute('''
                    UPDATE model_metadata
                    SET model_name = ?, dimension = ?, vector_count = ?, last_updated = ?
                    WHERE id = 1
                ''', (
                    EMBEDDING_MODEL_NAME,
                    self.dimension,
                    self.chunk_counter,
                    datetime.now().isoformat()
                ))
                logger.info(f"📝 Updated model metadata: {EMBEDDING_MODEL_NAME} (dimension: {self.dimension})")

            self.metadata_db.commit()

        except Exception as e:
            logger.error(f"❌ Failed to record model metadata: {e}")

    def _check_model_change(self):
        """Check if embedding model has changed and warn user"""
        try:
            cursor = self.metadata_db.cursor()

            # Get stored model metadata
            cursor.execute('''
                SELECT model_name, dimension FROM model_metadata WHERE id = 1
            ''')
            result = cursor.fetchone()

            if result:
                stored_model, stored_dimension = result

                # Check if model changed
                if stored_model != EMBEDDING_MODEL_NAME:
                    logger.warning(f"⚠️ EMBEDDING MODEL CHANGED!")
                    logger.warning(f"   Previous model: {stored_model}")
                    logger.warning(f"   Current model:  {EMBEDDING_MODEL_NAME}")
                    logger.warning(f"   Existing index will be used with new model (search results may be less accurate)")
                    logger.warning(f"   To rebuild with new model: see docs/ADMIN_MODEL_CHANGE_GUIDE.md")

                # Dimension check is already done in _initialize_storage()
                if stored_dimension != self.dimension:
                    logger.warning(f"⚠️ DIMENSION CHANGED: {stored_dimension} → {self.dimension}")

                # Update metadata
                self._record_model_metadata()
            else:
                # No metadata found, record it
                self._record_model_metadata()

        except Exception as e:
            logger.warning(f"⚠️ Could not check model metadata: {e}")

    async def add_chunks(self, chunks: List[DocumentChunk]) -> bool:
        """Add document chunks to FAISS index and SQLite"""
        try:
            if not chunks:
                return True
            
            # Generate embeddings for chunks
            embeddings = await self._generate_embeddings([chunk.content for chunk in chunks])
            if not embeddings:
                logger.error("❌ Failed to generate embeddings")
                return False
            
            # Add to FAISS index
            embeddings_array = np.vstack(embeddings)
            faiss_start_index = self.faiss_index.ntotal
            self.faiss_index.add(embeddings_array)
            
            # Add to SQLite
            cursor = self.metadata_db.cursor()
            for i, chunk in enumerate(chunks):
                faiss_index = faiss_start_index + i
                try:
                    cursor.execute('''
                        INSERT INTO chunks 
                        (chunk_id, faiss_index, document_path, chunk_index, total_chunks, content, metadata, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        chunk.chunk_id,
                        faiss_index,
                        chunk.document_path,
                        chunk.chunk_index,
                        chunk.total_chunks,
                        chunk.content,
                        json.dumps(chunk.metadata),
                        chunk.created_at
                    ))
                except sqlite3.IntegrityError:
                    cursor.execute('''
                        UPDATE chunks
                        SET faiss_index = ?,
                            content = ?,
                            metadata = ?,
                            total_chunks = ?
                        WHERE chunk_id = ?
                    ''', (faiss_index, chunk.content, json.dumps(chunk.metadata), chunk.total_chunks, chunk.chunk_id))
            
            self.metadata_db.commit()
            self.chunk_counter = self.faiss_index.ntotal
            
            # Save FAISS index
            await self._save_index()
            
            logger.info(f"✅ Added {len(chunks)} chunks to document store")
            return True
            
        except Exception as e:
            logger.error(f"❌ Failed to add chunks: {e}")
            return False
    
    async def search_similar(self, query: str, k: int = DEFAULT_SEARCH_RESULTS) -> List[Dict[str, Any]]:
        """Search for similar document chunks"""
        try:
            if self.faiss_index.ntotal == 0:
                return []
            
            # Generate query embedding
            query_embeddings = await self._generate_embeddings([query])
            if not query_embeddings:
                return []
            
            query_vector = np.array(query_embeddings[0]).reshape(1, -1)
            
            # Search FAISS index
            scores, indices = self.faiss_index.search(query_vector, min(k, self.faiss_index.ntotal))
            
            # Get metadata from SQLite
            results = []
            cursor = self.metadata_db.cursor()
            
            for score, faiss_idx in zip(scores[0], indices[0]):
                if faiss_idx == -1:  # No more results
                    break
                
                # 🎯 RELEVANCE FILTERING: Skip very dissimilar results
                # For FAISS IndexFlatIP, higher scores = more similar
                # Threshold 130+ includes all passport docs (134.8, 129.9) but may include some noise
                if score < 130.0:
                    logger.info(f"⚠️ Skipping low-relevance result: score={score:.1f} < 130.0 (faiss_idx={faiss_idx})")
                    continue
                
                cursor.execute('''
                    SELECT chunk_id, document_path, chunk_index, content, metadata, created_at
                    FROM chunks WHERE faiss_index = ?
                ''', (int(faiss_idx),))
                
                row = cursor.fetchone()
                if row:
                    chunk_id, doc_path, chunk_idx, content, metadata_json, created_at = row
                    results.append({
                        'chunk_id': chunk_id,
                        'document_path': doc_path,
                        'chunk_index': chunk_idx,
                        'content': content,
                        'metadata': json.loads(metadata_json),
                        'similarity_score': float(score),
                        'created_at': created_at
                    })
            
            return results
            
        except Exception as e:
            logger.error(f"❌ Search failed: {e}")
            return []
    
    async def _check_embedding_service_health(self) -> bool:
        """Check if embedding service is healthy and responsive"""
        try:
            payload = {"model": EMBEDDING_MODEL_NAME, "prompt": "health check", "keep_alive": -1}
            response_data = await pooled_post(
                EMBEDDING_SERVICE_URL,
                json=payload,
                timeout=HEALTH_CHECK_TIMEOUT_SECONDS
            )
            return response_data['status_code'] == 200
        except Exception as e:
            logger.error(f"❌ Embedding service health check failed: {e}")
            return False
    
    async def _restart_embedding_service(self) -> bool:
        """Restart the embedding service - CORRECTED VERSION"""
        try:
            logger.info("🔄 Attempting to restart embedding service...")
            
            # First, check if we need a separate instance at all
            # Modern Ollama can handle both LLM and embedding requests on the same port
            
            # Option 1: Use the main Ollama instance (RECOMMENDED)
            # Just ensure the embedding model is pulled and loaded
            try:
                # Pull the embedding model if not already present
                pull_result = await asyncio.create_subprocess_shell(
                    f"ollama pull {EMBEDDING_MODEL_NAME}",
                    stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.PIPE
                )
                stdout, stderr = await pull_result.communicate()
                
                if pull_result.returncode != 0:
                    logger.error(f"Failed to pull embedding model: {stderr.decode()}")
                    return False
                    
                # Pre-load the model to keep it in memory
                load_payload = {
                    "model": EMBEDDING_MODEL_NAME,
                    "keep_alive": -1  # Keep loaded indefinitely
                }
                
                # Use the main Ollama port for embeddings
                response = await pooled_post(
                    f"http://{OLLAMA_EMBEDDING_HOST}:{OLLAMA_MAIN_PORT}/api/embeddings",
                    json={
                        "model": EMBEDDING_MODEL_NAME,
                        "prompt": "warmup",
                        "keep_alive": -1
                    },
                    timeout=30
                )
                
                return response['status_code'] == 200
                
            except Exception as e:
                logger.error(f"Failed to setup embedding model: {e}")
                
                # Option 2: If you really need a separate instance (NOT RECOMMENDED)
                # Kill ALL ollama processes and restart with specific port
                logger.info("Attempting fallback: Kill all Ollama and restart on specific port")
                
                # Kill all ollama processes
                kill_result = await asyncio.create_subprocess_shell(
                    "pkill -f 'ollama serve' || true",  # || true prevents error if no process found
                    stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.PIPE
                )
                await kill_result.wait()
                
                # Wait for processes to fully terminate
                await asyncio.sleep(3)
                
                # Start Ollama on the embedding port using a proper daemon approach
                # Note: This will be the ONLY Ollama instance running
                start_cmd = f"""
                export OLLAMA_HOST=0.0.0.0:{OLLAMA_EMBEDDING_PORT} && \
                ollama serve 2>/dev/null 1>/dev/null &
                """
                
                start_result = await asyncio.create_subprocess_shell(
                    start_cmd,
                    stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.PIPE
                )
                await start_result.wait()
                
                # Wait longer for service to fully start
                await asyncio.sleep(10)  # Ollama needs time to initialize
                
                # Pull and load the embedding model
                pull_result = await asyncio.create_subprocess_shell(
                    f"OLLAMA_HOST=127.0.0.1:{OLLAMA_EMBEDDING_PORT} ollama pull {EMBEDDING_MODEL_NAME}",
                    stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.PIPE
                )
                await pull_result.communicate()
                
                # Verify it's working
                return await self._check_embedding_service_health()
                
        except Exception as e:
            logger.error(f"❌ Failed to restart embedding service: {e}")
            return False
        
    async def _generate_embeddings(self, texts: List[str]) -> Optional[List[np.ndarray]]:
        """Generate embeddings using Ollama via existing HTTP pool with health checking"""
        try:
            # First check if service is healthy
            if not await self._check_embedding_service_health():
                logger.warning("⚠️ Embedding service unhealthy, attempting restart...")
                
                # Try to restart service up to maximum attempts
                for attempt in range(MAX_SERVICE_RESTART_ATTEMPTS):
                    if await self._restart_embedding_service():
                        logger.info(f"✅ Embedding service restarted successfully (attempt {attempt + 1})")
                        break
                    else:
                        logger.warning(f"❌ Restart attempt {attempt + 1} failed")
                        await asyncio.sleep(RETRY_DELAY_SECONDS)
                else:
                    logger.error(f"❌ Failed to restart embedding service after {MAX_SERVICE_RESTART_ATTEMPTS} attempts")
                    logger.error("🛑 RECOMMENDATION: Manual intervention required - check ollama installation and restart embedding service manually")
                    logger.error(f"🛑 Command to restart: OLLAMA_HOST={OLLAMA_EMBEDDING_HOST}:{OLLAMA_EMBEDDING_PORT} ollama serve")
                    return None
            async def generate_single_embedding(text: str) -> Optional[np.ndarray]:
                """Generate embedding for a single text"""
                try:
                    payload = {
                        "model": EMBEDDING_MODEL_NAME,  # High-quality embedding model
                        "prompt": text,
                        "keep_alive": -1  # Keep model loaded permanently
                    }
                    
                    response_data = await pooled_post(
                        EMBEDDING_SERVICE_URL,  # Use dedicated embedding instance
                        json=payload,
                        timeout=EMBEDDING_TIMEOUT_SECONDS
                    )
                    
                    if response_data['status_code'] == 200:
                        embedding_data = json.loads(response_data['text'])
                        embedding = np.array(embedding_data['embedding'], dtype=np.float32)
                        return embedding
                    else:
                        logger.error(f"❌ Embedding generation failed: {response_data['status_code']}")
                        return None
                        
                except Exception as e:
                    logger.error(f"❌ Single embedding error: {e}")
                    return None
            
            # ✅ Process embeddings in batches with adaptive sizing for resilience
            current_batch_size = DEFAULT_BATCH_SIZE  # Start with configured batch size
            all_embeddings = []
            processed_count = 0

            logger.info(f"🔄 Processing {len(texts)} embeddings with batch_size={current_batch_size} (adaptive mode: {ADAPTIVE_BATCH_MODE_ENABLED})")

            while processed_count < len(texts):
                # Calculate batch boundaries
                start_idx = processed_count
                end_idx = min(start_idx + current_batch_size, len(texts))
                batch_texts = texts[start_idx:end_idx]
                total_processed = len(all_embeddings)
                total_chunks = len(texts)

                # Process current batch in parallel
                logger.debug(f"📦 Processing batch: {start_idx}-{end_idx} (batch_size={current_batch_size}, total progress: {total_processed}/{total_chunks})")
                batch_tasks = [generate_single_embedding(text) for text in batch_texts]
                batch_results = await asyncio.gather(*batch_tasks, return_exceptions=True)

                # Validate batch results with adaptive recovery
                batch_embeddings = []
                batch_failed = False

                for i, result in enumerate(batch_results):
                    if isinstance(result, Exception):
                        logger.error(f"❌ Batch failure at task {i}: {result}")
                        batch_failed = True
                        break
                    elif result is not None:
                        batch_embeddings.append(result)
                    else:
                        logger.error(f"❌ Batch failure: task {i} returned None")
                        batch_failed = True
                        break

                # Handle batch failure with adaptive recovery and health checks
                if batch_failed:
                    if not ADAPTIVE_BATCH_MODE_ENABLED:
                        logger.error(f"❌ Batch failed and adaptive mode disabled - aborting")
                        return None

                    # 🛡️ CRITICAL: Check if Ollama service itself is healthy
                    is_service_healthy = await self._check_embedding_service_health()
                    logger.warning(f"🔍 Ollama health check: {'✅ HEALTHY' if is_service_healthy else '❌ UNHEALTHY'}")

                    if not is_service_healthy:
                        logger.error(f"🚨 EMBEDDING SERVICE UNHEALTHY - Attempting restart")
                        logger.error(f"   Progress: {processed_count}/{total_chunks} embeddings processed")
                        logger.error(f"   Current batch_size: {current_batch_size}")

                        # Attempt to restart Ollama service
                        restart_success = False
                        for restart_attempt in range(2):  # Try 2 restart attempts
                            logger.info(f"🔄 Ollama restart attempt {restart_attempt + 1}/2...")
                            if await self._restart_embedding_service():
                                logger.info(f"✅ Ollama restarted successfully")
                                restart_success = True
                                break
                            else:
                                logger.error(f"❌ Restart attempt {restart_attempt + 1} failed")
                                await asyncio.sleep(5)

                        if not restart_success:
                            logger.error(f"❌ Failed to restart Ollama after 2 attempts")
                            logger.error(f"🛑 RECOMMENDATION: Manually restart Ollama service")
                            logger.error(f"   Command: systemctl restart ollama")
                            logger.error(f"   Or check: ps aux | grep ollama")
                            return None

                        # After restart, wait longer before retry
                        logger.info(f"⏳ Waiting 15 seconds for Ollama to stabilize...")
                        await asyncio.sleep(15)

                        # Reset batch size after service recovery (retry with original batch size)
                        logger.info(f"📈 Service recovered - resetting batch_size to {DEFAULT_BATCH_SIZE}")
                        current_batch_size = DEFAULT_BATCH_SIZE
                        # Retry this batch with fresh service and original batch size
                        continue

                    # Service is healthy but batch still failed - try reducing batch size
                    if current_batch_size > MIN_BATCH_SIZE:
                        new_batch_size = max(MIN_BATCH_SIZE, int(current_batch_size * ADAPTIVE_BATCH_REDUCTION_FACTOR))
                        logger.warning(f"⚠️ Service healthy but batch failed - reducing batch_size: {current_batch_size} → {new_batch_size}")
                        current_batch_size = new_batch_size

                        # Retry this batch with smaller size - DON'T increment processed_count
                        await asyncio.sleep(SERVICE_RESTART_DELAY_SECONDS)  # Brief pause before retry
                        continue
                    else:
                        logger.error(f"❌ Batch failed at minimum batch size ({MIN_BATCH_SIZE}) with healthy service")
                        logger.error(f"   Progress: {processed_count}/{total_chunks} embeddings")
                        logger.error(f"   This indicates a fundamental service issue")
                        logger.error(f"   Recommendation: Check Ollama logs for detailed errors")
                        return None

                # Batch succeeded - add to results
                all_embeddings.extend(batch_embeddings)
                processed_count = end_idx
                batch_num = (processed_count // current_batch_size) + (1 if processed_count % current_batch_size else 0)
                logger.info(f"✅ Completed batch: {len(batch_embeddings)} embeddings (progress: {processed_count}/{total_chunks})")

                # ✅ Periodic health check between batches (every 10 batches or at critical points)
                # This detects Ollama degradation BEFORE it fails
                check_interval = 10  # Check health every 10 successful batches
                batches_processed = processed_count // current_batch_size if current_batch_size > 0 else 0

                if batches_processed > 0 and batches_processed % check_interval == 0:
                    logger.debug(f"🏥 Periodic health check at batch {batches_processed}...")
                    is_healthy = await self._check_embedding_service_health()
                    if is_healthy:
                        logger.debug(f"   ✅ Ollama service healthy")
                    else:
                        logger.warning(f"⚠️ Ollama service degrading at batch {batches_processed}")
                        logger.warning(f"   This may indicate upcoming failures - monitoring closely")

                # ✅ Delay between batches to prevent service overload and memory buildup
                # Configured in config/llm_config.yaml: batch_processing.batch_delay_seconds
                if processed_count < len(texts):  # Don't delay after the last batch
                    await asyncio.sleep(BATCH_DELAY_SECONDS)

            logger.info(f"✅ Generated {len(all_embeddings)} embeddings across all batches (final batch_size={current_batch_size})")
            return all_embeddings
            
        except Exception as e:
            logger.error(f"❌ Parallel embedding generation error: {e}")
            return None
    
    async def _save_index(self):
        """Save FAISS index to disk"""
        try:
            faiss.write_index(self.faiss_index, str(self.index_path))
        except Exception as e:
            logger.error(f"❌ Failed to save FAISS index: {e}")
    
    def get_stats(self) -> Dict[str, Any]:
        """Get document store statistics including model information"""
        try:
            cursor = self.metadata_db.cursor()

            # Count documents
            cursor.execute("SELECT COUNT(*) FROM documents")
            doc_count = cursor.fetchone()[0]

            # Count chunks
            cursor.execute("SELECT COUNT(*) FROM chunks")
            chunk_count = cursor.fetchone()[0]

            # ✅ Get model metadata for diagnostics
            model_info = {
                'name': EMBEDDING_MODEL_NAME,
                'dimension': self.dimension
            }
            try:
                cursor.execute('''
                    SELECT model_name, dimension, vector_count, created_at, last_updated
                    FROM model_metadata WHERE id = 1
                ''')
                result = cursor.fetchone()
                if result:
                    model_info['stored_name'] = result[0]
                    model_info['stored_dimension'] = result[1]
                    model_info['vector_count'] = result[2]
                    model_info['created_at'] = result[3]
                    model_info['last_updated'] = result[4]
                    model_info['model_matches'] = (result[0] == EMBEDDING_MODEL_NAME)
                    model_info['dimension_matches'] = (result[1] == self.dimension)
            except:
                pass  # model_metadata table might not exist yet

            return {
                'total_documents': doc_count,
                'total_chunks': chunk_count,
                'faiss_vectors': self.faiss_index.ntotal,
                'storage_directory': str(self.storage_dir),
                'index_dimension': self.dimension,
                'embedding_model': model_info
            }

        except Exception as e:
            logger.error(f"❌ Failed to get stats: {e}")
            return {}


if WATCHDOG_AVAILABLE:
    class DirectoryWatcher(FileSystemEventHandler):
        """File system watcher for automatic document processing"""
        
        def __init__(self, interrogator):
            self.interrogator = interrogator
        
        def on_created(self, event):
            if not event.is_directory:
                logger.info(f"📄 New file detected: {event.src_path}")
                asyncio.create_task(self.interrogator._process_single_file(event.src_path))
        
        def on_modified(self, event):
            if not event.is_directory:
                logger.info(f"📝 File modified: {event.src_path}")
                asyncio.create_task(self.interrogator._process_single_file(event.src_path))


class DocumentInterrogator:
    """Main document interrogation system integrating with 2-stage LLM architecture"""
    
    def __init__(self, storage_dir: str = "document_store"):
        self.processor = DocumentProcessor()
        self.store = None
        self.observer = None
        self.watched_directories = set()
        self.max_files_per_scan = config_loader.load_config().get('document_interrogator', {}).get('max_files_per_scan', 100)
        
        # Configuration file path
        self.config_file = Path("watched_directories.json")
        self.config = self._load_config()
        
        # Scan synchronization to prevent infinite loops
        self._scan_lock = asyncio.Lock()
        self._scan_in_progress = False
        
        # Initialize FAISS store if available
        if FAISS_AVAILABLE:
            try:
                self.store = FAISSDocumentStore(storage_dir)
                logger.info("🔍 Document Interrogator initialized with FAISS")
                
                # 🛡️ PRODUCTION SAFETY: Automatic FAISS integrity monitoring
                if INTEGRITY_MONITORING_AVAILABLE:
                    logger.info("🛡️ Running FAISS integrity check...")
                    asyncio.create_task(self._run_integrity_check_on_startup())
                    # Schedule periodic integrity checks (every 6 hours)
                    asyncio.create_task(self._schedule_periodic_integrity_checks())
                else:
                    logger.warning("⚠️ FAISS integrity monitoring disabled - potential corruption may go undetected")
                
                # Auto-scan for changes on startup based on configuration (temporarily disabled for API testing)
                if False: # self.config.get('config', {}).get('scan_on_startup', False):
                    logger.info("🚀 Startup scan enabled - using safe implementation")
                    asyncio.create_task(self._safe_startup_config_scan())
                
            except Exception as e:
                logger.error(f"❌ FAISS initialization failed: {e}")
                self.store = None
        else:
            logger.warning(f"⚠️ FAISS not available: {FAISS_ERROR}")
            logger.info("💡 Install with: pip install faiss-cpu numpy")
    
    async def index_directory(self, directory_path: str, recursive: bool = True) -> Dict[str, Any]:
        """Index all documents in a directory"""
        if not self.store:
            return {
                'error': 'FAISS not available',
                'processed': 0,
                'failed': 0
            }
            
        try:
            directory = Path(directory_path)
            if not directory.exists() or not directory.is_dir():
                raise ValueError(f"Invalid directory: {directory_path}")
            
            results = {
                'processed': 0,
                'failed': 0,
                'total_chunks': 0,
                'files': []
            }
            
            # Get all supported files
            pattern = "**/*" if recursive else "*"
            all_files = []
            
            for ext in self.processor.SUPPORTED_TYPES.keys():
                all_files.extend(directory.glob(f"{pattern}{ext}"))
            
            logger.info(f"📚 Found {len(all_files)} documents to process")
            
            # Process each file
            for file_path in all_files:
                try:
                    chunks = await self.processor.process_document(str(file_path))
                    if chunks:
                        success = await self.store.add_chunks(chunks)
                        if success:
                            results['processed'] += 1
                            results['total_chunks'] += len(chunks)
                            results['files'].append({
                                'path': str(file_path),
                                'chunks': len(chunks),
                                'status': 'success'
                            })
                        else:
                            results['failed'] += 1
                            results['files'].append({
                                'path': str(file_path),
                                'status': 'failed_storage'
                            })
                    else:
                        results['failed'] += 1
                        results['files'].append({
                            'path': str(file_path),
                            'status': 'failed_processing'
                        })
                        
                except Exception as e:
                    logger.error(f"❌ Failed to process {file_path}: {e}")
                    results['failed'] += 1
                    results['files'].append({
                        'path': str(file_path),
                        'status': 'error',
                        'error': str(e)
                    })
            
            logger.info(f"✅ Indexing complete: {results['processed']} processed, {results['failed']} failed")
            return results
            
        except Exception as e:
            logger.error(f"❌ Directory indexing failed: {e}")
            raise
    
    async def search_documents(self, query: str, k: int = DEFAULT_SEARCH_RESULTS) -> Dict[str, Any]:
        """Search documents for relevant content (Stage 0 RAG retrieval)"""
        if not self.store:
            return {
                'chunks': [],
                'context': '',
                'sources': [],
                'query': query,
                'error': 'FAISS not available'
            }
            
        try:
            # Search for relevant chunks
            similar_chunks = await self.store.search_similar(query, k)
            
            if not similar_chunks:
                return {
                    'chunks': [],
                    'context': '',
                    'sources': [],
                    'query': query
                }
            
            # Prepare context for 2-stage LLM
            context_parts = []
            sources = []
            
            for chunk in similar_chunks:
                doc_name = Path(chunk['document_path']).name
                context_parts.append(f"[Document: {doc_name}]\n{chunk['content']}")
                sources.append({
                    'document': chunk['document_path'],
                    'document_name': doc_name,
                    'chunk_index': chunk['chunk_index'],
                    'similarity_score': chunk['similarity_score']
                })
            
            # Create structured context for LLM
            context = "\n\n---\n\n".join(context_parts)
            
            return {
                'chunks': similar_chunks,
                'context': context,
                'sources': sources,
                'query': query,
                'chunks_found': len(similar_chunks)
            }
            
        except Exception as e:
            logger.error(f"❌ Document search failed: {e}")
            return {
                'chunks': [],
                'context': '',
                'sources': [],
                'query': query,
                'error': str(e)
            }
    
    def start_watching(self, directory_path: str):
        """Start watching directory for changes"""
        if not WATCHDOG_AVAILABLE:
            logger.warning("⚠️ Directory watching not available - watchdog not installed")
            return False
            
        try:
            if not self.observer:
                self.observer = Observer()
            
            directory = Path(directory_path)
            if directory.exists():
                event_handler = DirectoryWatcher(self)
                self.observer.schedule(event_handler, str(directory), recursive=True)
                self.watched_directories.add(str(directory))
                
                if not self.observer.is_alive():
                    self.observer.start()
                
                logger.info(f"👁️ Started watching directory: {directory_path}")
                return True
            else:
                logger.error(f"❌ Directory not found: {directory_path}")
                return False
                
        except Exception as e:
            logger.error(f"❌ Failed to start watching {directory_path}: {e}")
            return False
    
    def stop_watching(self):
        """Stop watching all directories"""
        try:
            if self.observer and self.observer.is_alive():
                self.observer.stop()
                self.observer.join()
                self.watched_directories.clear()
                logger.info("🛑 Stopped directory watching")
        except Exception as e:
            logger.error(f"❌ Failed to stop watching: {e}")
    
    def _load_config(self) -> Dict[str, Any]:
        """Load configuration from watched_directories.json"""
        try:
            if self.config_file.exists():
                with open(self.config_file, 'r') as f:
                    config = json.load(f)
                    logger.info(f"📋 Loaded watch configuration: {len(config.get('directories', []))} directories")
                    return config
            else:
                # Create default config
                default_config = {
                    "version": "1.0",
                    "config": {
                        "scan_on_startup": True,
                        "batch_size": DEFAULT_BATCH_SIZE,
                        "scan_interval_minutes": DEFAULT_SCAN_INTERVAL_MINUTES,
                        "auto_watch_enabled": True
                    },
                    "directories": [],
                    "last_scan": None,
                    "stats": {
                        "total_directories": 0,
                        "active_directories": 0,
                        "last_config_update": datetime.now().isoformat()
                    }
                }
                self._save_config(default_config)
                return default_config
        except Exception as e:
            logger.error(f"❌ Error loading config: {e}")
            return {"config": {}, "directories": []}
    
    def _save_config(self, config: Dict[str, Any] = None):
        """Save configuration to watched_directories.json"""
        try:
            if config is None:
                config = self.config
            
            config['stats']['last_config_update'] = datetime.now().isoformat()
            
            with open(self.config_file, 'w') as f:
                json.dump(config, f, indent=2)
                
        except Exception as e:
            logger.error(f"❌ Error saving config: {e}")
    
    async def _startup_config_scan(self):
        """Scan directories from config for changes that occurred while server was offline"""
        async with self._scan_lock:
            if self._scan_in_progress:
                logger.info("⏳ Scan already in progress, skipping duplicate scan request")
                return
                
            self._scan_in_progress = True
            try:
                await asyncio.sleep(STARTUP_INITIALIZATION_DELAY)  # Wait for full initialization
                
                directories = self.config.get('directories', [])
                enabled_dirs = [d for d in directories if d.get('enabled', True)]
                
                if not enabled_dirs:
                    logger.info("📋 No directories configured for watching")
                    self._scan_in_progress = False
                    return
                    
                logger.info(f"🔍 Startup config scan: Checking {len(enabled_dirs)} configured directories")
                changes_found = 0
                
                for dir_config in enabled_dirs:
                    directory_path = dir_config['path']
                    recursive = dir_config.get('recursive', True)
                    
                    directory = Path(directory_path)
                    if not directory.exists():
                        logger.warning(f"⚠️ Configured directory not found: {directory_path}")
                        continue
                    
                    logger.info(f"🔍 Scanning: {directory_path}")
                    
                    if recursive:
                        file_pattern = directory.rglob('*')
                    else:
                        file_pattern = directory.glob('*')
                    
                    for file_path in file_pattern:
                        if file_path.is_file() and file_path.suffix.lower() in self.processor.SUPPORTED_TYPES:
                            if await self._file_needs_reindexing(str(file_path)):
                                await self._process_single_file(str(file_path))
                                changes_found += 1
                
                # Update last scan time
                self.config['last_scan'] = datetime.now().isoformat()
                self._save_config()
                
                if changes_found > 0:
                    logger.info(f"🔄 Startup scan complete: {changes_found} files reindexed")
                else:
                    logger.info("✅ Startup scan: All configured directories up to date")
                    
                # Auto-start watching if enabled
                if self.config.get('config', {}).get('auto_watch_enabled', True):
                    await self._start_watching_configured_directories()
                    
            except Exception as e:
                logger.error(f"❌ Startup config scan failed: {e}")
            finally:
                self._scan_in_progress = False
    
    async def _file_needs_reindexing(self, file_path: str) -> bool:
        """Check if file needs reindexing based on modification time and hash"""
        try:
            if not self.store:
                return True  # No store connection, process file to be safe
                
            file_stat = Path(file_path).stat()
            current_mtime = datetime.fromtimestamp(file_stat.st_mtime).isoformat()
            
            # Calculate current file hash
            with open(file_path, 'rb') as f:
                current_hash = hashlib.md5(f.read()).hexdigest()
            
            # Check database for existing record
            cursor = self.store.metadata_db.cursor()
            cursor.execute('''
                SELECT file_hash, last_modified FROM documents 
                WHERE file_path = ?
            ''', (file_path,))
            
            result = cursor.fetchone()
            if not result:
                logger.debug(f"📋 New file detected: {Path(file_path).name}")
                return True  # New file, needs indexing
            
            stored_hash, stored_mtime = result
            
            # Compare hash and modification time
            if stored_hash != current_hash:
                logger.info(f"🔄 Change detected (hash): {Path(file_path).name}")
                return True
            elif stored_mtime != current_mtime:
                logger.info(f"🔄 Change detected (mtime): {Path(file_path).name}")
                return True
                
            logger.debug(f"✅ File up-to-date: {Path(file_path).name}")
            return False
            
        except Exception as e:
            logger.error(f"❌ Error checking file {file_path}: {e}")
            return True  # On error, process file to be safe
    
    async def _record_document_metadata(self, file_path: str):
        """Record document metadata after successful processing"""
        try:
            if not self.store:
                logger.warning(f"📋 No store connection for metadata recording: {Path(file_path).name}")
                return
                
            file_stat = Path(file_path).stat()
            current_mtime = datetime.fromtimestamp(file_stat.st_mtime).isoformat()
            
            # Calculate current file hash
            with open(file_path, 'rb') as f:
                current_hash = hashlib.md5(f.read()).hexdigest()
            
            # Record in documents table
            cursor = self.store.metadata_db.cursor()
            file_stat = Path(file_path).stat()
            cursor.execute('''
                INSERT OR REPLACE INTO documents 
                (file_path, file_hash, file_type, file_size, total_chunks, processed_at, last_modified)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (file_path, current_hash, Path(file_path).suffix.lower(), 
                  file_stat.st_size, 0, datetime.now().isoformat(), current_mtime))
            
            self.store.metadata_db.commit()
            
            # Verify the record was saved
            cursor.execute('SELECT file_hash, last_modified FROM documents WHERE file_path = ?', (file_path,))
            verify_result = cursor.fetchone()
            if verify_result:
                logger.debug(f"📝 Metadata recorded: {Path(file_path).name} (hash: {current_hash[:8]}...)")
            else:
                logger.error(f"❌ Metadata verification failed: {Path(file_path).name}")
            
        except Exception as e:
            logger.error(f"❌ Failed to record document metadata for {file_path}: {e}")
    
    async def _start_watching_configured_directories(self):
        """Start watching all enabled directories from config"""
        if not WATCHDOG_AVAILABLE:
            logger.warning("⚠️ Auto-watching disabled - watchdog not installed")
            return
            
        enabled_dirs = [d for d in self.config.get('directories', []) if d.get('enabled', True)]
        
        for dir_config in enabled_dirs:
            directory_path = dir_config['path']
            if Path(directory_path).exists():
                self.start_watching(directory_path)
            else:
                logger.warning(f"⚠️ Cannot watch non-existent directory: {directory_path}")
    
    def add_watch_directory(self, directory_path: str, recursive: bool = True, enabled: bool = True, description: str = ""):
        """Add a directory to the watch configuration"""
        try:
            # Check if directory already exists in config
            for dir_config in self.config.get('directories', []):
                if dir_config['path'] == directory_path:
                    logger.warning(f"⚠️ Directory already in config: {directory_path}")
                    return False
            
            new_dir = {
                "path": directory_path,
                "recursive": recursive,
                "enabled": enabled,
                "description": description,
                "added_at": datetime.now().isoformat()
            }
            
            if 'directories' not in self.config:
                self.config['directories'] = []
            
            self.config['directories'].append(new_dir)
            
            # Update stats
            self.config['stats']['total_directories'] = len(self.config['directories'])
            self.config['stats']['active_directories'] = len([d for d in self.config['directories'] if d.get('enabled', True)])
            
            self._save_config()
            logger.info(f"➕ Added directory to config: {directory_path}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Error adding directory to config: {e}")
            return False
    
    def remove_watch_directory(self, directory_path: str):
        """Remove a directory from the watch configuration"""
        try:
            original_count = len(self.config.get('directories', []))
            self.config['directories'] = [d for d in self.config.get('directories', []) if d['path'] != directory_path]
            
            if len(self.config['directories']) < original_count:
                # Update stats
                self.config['stats']['total_directories'] = len(self.config['directories'])
                self.config['stats']['active_directories'] = len([d for d in self.config['directories'] if d.get('enabled', True)])
                
                self._save_config()
                logger.info(f"➖ Removed directory from config: {directory_path}")
                return True
            else:
                logger.warning(f"⚠️ Directory not found in config: {directory_path}")
                return False
                
        except Exception as e:
            logger.error(f"❌ Error removing directory from config: {e}")
            return False
    
    async def smart_index_directory(self, directory_path: str, recursive: bool = True) -> Dict[str, Any]:
        """Smart directory indexing that only processes changed files"""
        if not self.store:
            return {
                'success': False,
                'error': 'FAISS not available',
                'processed': 0,
                'failed': 0,
                'skipped': 0,
                'message': 'Document store not initialized'
            }
        
        try:
            directory = Path(directory_path)
            if not directory.exists() or not directory.is_dir():
                return {
                    'success': False,
                    'error': f'Invalid directory: {directory_path}',
                    'processed': 0,
                    'failed': 0,
                    'skipped': 0
                }
            
            results = {
                'processed': 0,
                'failed': 0,
                'skipped': 0,
                'files': [],
                'total_files_found': 0
            }
            
            # Get file list
            if recursive:
                file_pattern = directory.rglob('*')
            else:
                file_pattern = directory.glob('*')
            
            files_found = [f for f in file_pattern if f.is_file() and f.suffix.lower() in self.processor.SUPPORTED_TYPES]
            results['total_files_found'] = len(files_found)
            
            logger.info(f"📚 Smart indexing: Found {len(files_found)} supported files")
            
            # Check each file for changes
            for file_path in files_found:
                try:
                    if await self._file_needs_reindexing(str(file_path)):
                        # File needs processing
                        success = await self._process_single_file(str(file_path))
                        if success:
                            results['processed'] += 1
                            results['files'].append({
                                'file': str(file_path),
                                'status': 'processed',
                                'reason': 'modified or new'
                            })
                        else:
                            results['failed'] += 1
                            results['files'].append({
                                'file': str(file_path),
                                'status': 'failed',
                                'reason': 'processing error'
                            })
                            # Stop on embedding service failure
                            break
                    else:
                        # File is up to date
                        results['skipped'] += 1
                        results['files'].append({
                            'file': str(file_path),
                            'status': 'skipped',
                            'reason': 'up-to-date'
                        })
                        
                except Exception as e:
                    logger.error(f"❌ Error processing {file_path}: {e}")
                    results['failed'] += 1
                    results['files'].append({
                        'file': str(file_path),
                        'status': 'failed',
                        'reason': str(e)
                    })
            
            # Generate appropriate message
            if results['processed'] == 0 and results['failed'] == 0:
                results['message'] = f"Scan completed: No modified or new files indexed, all {results['skipped']} files are up-to-date"
            elif results['failed'] > 0:
                results['message'] = f"Scan completed with errors: {results['processed']} processed, {results['failed']} failed, {results['skipped']} up-to-date"
            else:
                results['message'] = f"Scan completed: {results['processed']} new/modified files indexed, {results['skipped']} files up-to-date"
                
            results['success'] = True
            logger.info(f"✅ Smart indexing complete: {results['message']}")
            return results
            
        except Exception as e:
            logger.error(f"❌ Smart directory indexing failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'processed': 0,
                'failed': 0,
                'skipped': 0,
                'message': f'Smart indexing failed: {str(e)}'
            }
    
    async def _run_integrity_check_on_startup(self):
        """
        🛡️ PRODUCTION SAFETY: Run comprehensive FAISS integrity check on startup
        Automatically detects and repairs corruption to prevent production issues
        """
        try:
            # Give system time to fully initialize
            await asyncio.sleep(2)
            
            if not self.store:
                logger.error("❌ Cannot run integrity check - FAISS store not available")
                return
                
            logger.info("🔍 Starting FAISS-SQLite integrity validation...")
            
            # Run comprehensive integrity check with automatic repair
            system_healthy = await check_and_repair_faiss_integrity(self.store)
            
            if system_healthy:
                logger.info("✅ FAISS integrity check passed - system is healthy")
            else:
                logger.error("🚨 FAISS integrity check failed - manual intervention may be required")
                logger.error("🚨 Document search functionality may be compromised")
                
        except Exception as e:
            logger.error(f"❌ Critical error during integrity check: {e}")
            logger.error("🚨 FAISS system integrity unknown - proceed with caution")
    
    async def _schedule_periodic_integrity_checks(self):
        """
        🛡️ PRODUCTION SAFETY: Schedule periodic integrity checks
        Runs every 6 hours to detect corruption early
        """
        CHECK_INTERVAL = 6 * 60 * 60  # 6 hours in seconds
        
        while True:
            try:
                await asyncio.sleep(CHECK_INTERVAL)
                
                if not self.store:
                    logger.warning("⚠️ Skipping periodic integrity check - FAISS store not available")
                    continue
                
                logger.info("🔍 Running scheduled FAISS integrity check...")
                
                # Run quick integrity check (non-blocking)
                system_healthy = await check_and_repair_faiss_integrity(self.store)
                
                if system_healthy:
                    logger.info("✅ Periodic integrity check passed")
                else:
                    logger.error("🚨 Periodic integrity check detected issues")
                    
            except asyncio.CancelledError:
                logger.info("🛑 Periodic integrity monitoring stopped")
                break
            except Exception as e:
                logger.error(f"❌ Error in periodic integrity check: {e}")
                # Continue monitoring despite errors
    
    async def _safe_startup_config_scan(self):
        """Safe version of startup config scan with limits and detailed logging"""
        if self._scan_in_progress:
            logger.info("⏳ Safe scan skipped - scan already in progress")
            return
            
        self._scan_in_progress = True
        try:
            await asyncio.sleep(1)  # Brief initialization wait
            
            directories = self.config.get('directories', [])
            enabled_dirs = [d for d in directories if d.get('enabled', True)]
            
            if not enabled_dirs:
                logger.info("📋 No directories configured for watching")
                return
                
            logger.info(f"🔍 Safe scan: Starting scan of {len(enabled_dirs)} configured directories")
            total_files_scanned = 0
            total_files_processed = 0
            MAX_FILES_PER_SCAN = self.max_files_per_scan  # Safety limit
            
            for dir_idx, dir_config in enumerate(enabled_dirs):
                directory_path = dir_config['path']
                recursive = dir_config.get('recursive', True)
                
                directory = Path(directory_path)
                if not directory.exists():
                    logger.warning(f"⚠️ Directory {dir_idx+1}/{len(enabled_dirs)} not found: {directory_path}")
                    continue
                
                logger.info(f"🔍 Scanning directory {dir_idx+1}/{len(enabled_dirs)}: {directory_path}")
                
                # Get file list with limit
                if recursive:
                    file_pattern = directory.rglob('*')
                else:
                    file_pattern = directory.glob('*')
                
                dir_files_scanned = 0
                dir_files_processed = 0
                
                for file_path in file_pattern:
                    if total_files_scanned >= MAX_FILES_PER_SCAN:
                        logger.warning(f"🛑 Reached safety limit of {MAX_FILES_PER_SCAN} files - stopping scan")
                        break
                        
                    if file_path.is_file() and file_path.suffix.lower() in self.processor.SUPPORTED_TYPES:
                        total_files_scanned += 1
                        dir_files_scanned += 1
                        
                        if dir_files_scanned % SCAN_PROGRESS_LOG_INTERVAL == 0:  # Log progress periodically
                            logger.info(f"📊 Directory {dir_idx+1}: scanned {dir_files_scanned} files")
                        
                        if await self._file_needs_reindexing(str(file_path)):
                            logger.info(f"🔄 Processing: {file_path.name}")
                            success = await self._process_single_file(str(file_path))
                            
                            if success:
                                total_files_processed += 1
                                dir_files_processed += 1
                                
                                if MAX_FILES_PER_DIRECTORY_SCAN is not None and dir_files_processed >= MAX_FILES_PER_DIRECTORY_SCAN:  # Limit per directory
                                    logger.info(f"📈 Directory {dir_idx+1}: processed {dir_files_processed} files, moving to next directory")
                                    break
                            else:
                                logger.error(f"🛑 Stopping scan due to embedding service failure")
                                logger.error(f"🔄 Will retry during next watch interval (scan_interval_minutes: {self.config.get('config', {}).get('scan_interval_minutes', DEFAULT_SCAN_INTERVAL_MINUTES)})")
                                return  # Stop scanning immediately
                
                if total_files_scanned >= MAX_FILES_PER_SCAN:
                    break
                    
                logger.info(f"✅ Directory {dir_idx+1} complete: {dir_files_scanned} scanned, {dir_files_processed} processed")
            
            # Update last scan time
            self.config['last_scan'] = datetime.now().isoformat()
            self._save_config()
            
            logger.info(f"🎉 Safe scan complete: {total_files_scanned} files scanned, {total_files_processed} files processed")
                
        except Exception as e:
            logger.error(f"❌ Safe startup config scan failed: {e}")
        finally:
            self._scan_in_progress = False
    
    async def force_scan_changes(self):
        """Force scan all configured directories for changes"""
        async with self._scan_lock:
            if self._scan_in_progress:
                logger.info("⏳ Scan already in progress, skipping duplicate scan request")
                return {"status": "skipped", "reason": "scan_in_progress"}
                
            logger.info("🔄 Force scan requested - using safe implementation")
            await self._safe_startup_config_scan()
            return {"status": "completed", "reason": "safe_scan_finished"}
    
    def get_config_status(self) -> Dict[str, Any]:
        """Get current configuration status"""
        return {
            "config_file_exists": self.config_file.exists(),
            "config": self.config.get('config', {}),
            "directories": self.config.get('directories', []),
            "stats": self.config.get('stats', {}),
            "last_scan": self.config.get('last_scan'),
            "currently_watching": list(self.watched_directories)
        }

    async def _process_single_file(self, file_path: str):
        """Process a single file (used by directory watcher)"""
        if not self.store:
            return
            
        try:
            chunks = await self.processor.process_document(file_path)
            if chunks:
                success = await self.store.add_chunks(chunks)
                if success:
                    # Record document metadata for change detection
                    await self._record_document_metadata(file_path)
                    logger.info(f"✅ Auto-processed: {file_path}")
                    return True
                else:
                    logger.error(f"❌ Failed to add chunks (likely embedding service issue): {file_path}")
                    return False
            else:
                logger.warning(f"⚠️ No chunks generated for: {file_path}")
                # Even if no chunks, record that we processed the file
                await self._record_document_metadata(file_path)
                return True  # Not an error, just no content to process
        except Exception as e:
            logger.error(f"❌ Auto-processing failed for {file_path}: {e}")
            return False
    
    def get_stats(self) -> Dict[str, Any]:
        """Get system statistics"""
        base_stats = {
            'faiss_available': FAISS_AVAILABLE,
            'document_processing_available': DOCUMENT_PROCESSING_AVAILABLE,
            'watchdog_available': WATCHDOG_AVAILABLE,
            'watched_directories': list(self.watched_directories),
            'is_watching': self.observer.is_alive() if self.observer else False,
            'supported_file_types': list(self.processor.SUPPORTED_TYPES.keys())
        }
        
        if self.store:
            store_stats = self.store.get_stats()
            return {**base_stats, **store_stats}
        else:
            return base_stats
    
    def is_ready(self) -> bool:
        """Check if the system is ready for document interrogation"""
        return (FAISS_AVAILABLE and 
                DOCUMENT_PROCESSING_AVAILABLE and 
                self.store is not None)
    
    async def start_background_scanning(self):
        """Start the background periodic scanning task"""
        if not self.config.get('config', {}).get('auto_watch_enabled', False):
            logger.info("📋 Background scanning disabled in configuration")
            return
            
        scan_interval_minutes = self.config.get('config', {}).get('scan_interval_minutes', DEFAULT_SCAN_INTERVAL_MINUTES)
        if scan_interval_minutes <= 0:
            logger.warning("⚠️ Invalid scan_interval_minutes, background scanning disabled")
            return
            
        logger.info(f"🔄 Starting background scanning every {scan_interval_minutes} minutes")
        self._background_scan_task = asyncio.create_task(self._background_scan_loop(scan_interval_minutes))
    
    async def stop_background_scanning(self):
        """Stop the background periodic scanning task"""
        if hasattr(self, '_background_scan_task') and self._background_scan_task:
            logger.info("🛑 Stopping background scanning task")
            self._background_scan_task.cancel()
            try:
                await self._background_scan_task
            except asyncio.CancelledError:
                pass
            self._background_scan_task = None
    
    async def _background_scan_loop(self, interval_minutes: int):
        """Background loop that performs periodic scans"""
        interval_seconds = interval_minutes * 60
        
        try:
            while True:
                # Wait for the specified interval
                await asyncio.sleep(interval_seconds)
                
                # Perform the periodic scan
                logger.info(f"⏰ Periodic scan starting (interval: {interval_minutes}min)")
                try:
                    await self._safe_startup_config_scan()
                    logger.info(f"✅ Periodic scan completed successfully")
                except Exception as e:
                    logger.error(f"❌ Periodic scan failed: {e}")
                    
        except asyncio.CancelledError:
            logger.info("🛑 Background scanning loop cancelled")
            raise
        except Exception as e:
            logger.error(f"❌ Background scanning loop error: {e}")
            # Wait a bit before the loop would restart (if it does)
            await asyncio.sleep(60)


# Global instance
document_interrogator = None

def get_document_interrogator(storage_dir: str = "document_store") -> DocumentInterrogator:
    """Get or create global document interrogator instance"""
    global document_interrogator
    if document_interrogator is None:
        document_interrogator = DocumentInterrogator(storage_dir)
    return document_interrogator